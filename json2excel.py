#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
JSON -> Excel (xlsx) Rückkonvertierung für das zuvor definierte JSON-Schema.
- Liest alle .json im Script-Verzeichnis
- Erzeugt je JSON eine .xlsx (gleicher Basisname) und .log.txt
- Rekonstruiert Sheets, Excel-Tabellen (inkl. Header & Daten), Merged-Cells, Sheet-Visibility
- Schreibt ausführliches Mapping-Log von JSON-Pfaden zu Excel-Zellen
"""

from __future__ import annotations
import json
import logging
import sys
from pathlib import Path
from typing import Any, Dict, List, Tuple
from datetime import datetime, date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import range_boundaries

# ----------------------------
# Utils & Logging
# ----------------------------

def script_dir() -> Path:
    try:
        return Path(__file__).resolve().parent
    except NameError:
        return Path.cwd().resolve()

def setup_logger(log_path: Path) -> logging.Logger:
    logger = logging.getLogger(log_path.stem)
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")
    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)
    sh = logging.StreamHandler(sys.stdout)
    sh.setFormatter(fmt)
    logger.addHandler(sh)
    return logger

def cell_addr(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"

def coerce_value(v: Any) -> Any:
    # Rück-Konvertierungen für einfache Typen (JSON → Excel)
    if isinstance(v, str):
        # ISO-Date/Datetime heuristisch: nur wenn eindeutig
        try:
            if len(v) >= 10 and v[4] == "-" and v[7] == "-":
                # datetime?
                try:
                    return datetime.fromisoformat(v)
                except Exception:
                    return date.fromisoformat(v)
        except Exception:
            pass
        return v
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, bool) or v is None:
        return v
    if isinstance(v, (datetime, date, Decimal)):
        return v
    # Fallback
    return str(v)

# ----------------------------
# Sheet-Rekonstruktion
# ----------------------------

def ensure_unique_table_name(existing: set, desired: str) -> str:
    base = desired if desired else "Table"
    name = base
    i = 1
    while name in existing:
        i += 1
        name = f"{base}_{i}"
    existing.add(name)
    return name

def write_merged_ranges(ws: Worksheet, merges: List[Dict[str, Any]], logger: logging.Logger) -> None:
    if not merges:
        return
    for m in merges:
        r = m.get("range")
        if not r:
            continue
        try:
            ws.merge_cells(r)
            logger.info(f"Merged: {ws.title}!{r}")
        except Exception as e:
            logger.warning(f"Merge übersprungen ({ws.title}!{r}): {e}")

def write_used_range(ws: Worksheet, used_range: Dict[str, Any], logger: logging.Logger) -> None:
    grid: List[List[Any]] = used_range.get("grid") or []
    # optional: dimensions/row_count/col_count werden nicht zwingend benötigt
    for r_idx, row_vals in enumerate(grid, start=1):
        for c_idx, v in enumerate(row_vals, start=1):
            addr = cell_addr(r_idx, c_idx)
            ws.cell(row=r_idx, column=c_idx, value=coerce_value(v))
            logger.info(f"Map: used_range.grid[{r_idx-1}][{c_idx-1}] -> {ws.title}!{addr}")
    # Merges (Sheet-weit oder used_range-spezifisch)
    merges = used_range.get("merged_cells") or []
    write_merged_ranges(ws, merges, logger)

def write_tables(ws: Worksheet, tables_json: List[Dict[str, Any]], logger: logging.Logger) -> None:
    """
    Schreibt eine Liste von Tabellen auf das Worksheet.
    Nutzt table['ref'] als Startbereich; passt die Endzeile dynamisch an row_count an.
    """
    existing_names = {t.displayName for t in ws._tables} if hasattr(ws, "_tables") else set()

    for t_idx, t in enumerate(tables_json):
        # Basis aus JSON
        headers = t.get("headers") or []
        rows = t.get("rows") or []
        ref = t.get("ref")  # z.B. "A1:D20"
        if not ref:
            # Fallback: lege ab Zeile 1, Spalte 1; Tabellen hintereinander mit Leerzeile
            start_row = ws.max_row + 2 if ws.max_row > 1 else 1
            start_col = 1
        else:
            min_col, min_row, max_col, max_row = range_boundaries(ref)
            start_row, start_col = min_row, min_col

        n_cols = max(len(headers), max((len(r.keys()) for r in rows), default=0))
        # Spaltenüberschriften schreiben
        for c in range(n_cols):
            header_val = headers[c] if c < len(headers) else f"Col_{c+1}"
            addr = cell_addr(start_row, start_col + c)
            ws.cell(row=start_row, column=start_col + c, value=header_val)
            logger.info(f"Map: tables[{t_idx}].headers[{c}] -> {ws.title}!{addr}")

        # Daten schreiben
        for r_i, row_obj in enumerate(rows, start=1):
            for c in range(n_cols):
                key = headers[c] if c < len(headers) else f"Col_{c+1}"
                val = row_obj.get(key, None)
                addr = cell_addr(start_row + r_i, start_col + c)
                ws.cell(row=start_row + r_i, column=start_col + c, value=coerce_value(val))
                logger.info(f"Map: tables[{t_idx}].rows[{r_i-1}]['{key}'] -> {ws.title}!{addr}")

        # Tabellengrenzen berechnen
        end_row = start_row + len(rows)
        end_col = start_col + n_cols - 1
        table_ref = f"{cell_addr(start_row, start_col)}:{cell_addr(end_row, end_col)}"

        # Tabellennamen (eindeutig)
        desired_name = t.get("name") or f"Table{t_idx+1}"
        display_name = ensure_unique_table_name(existing_names, desired_name)

        # Table hinzufügen
        try:
            tbl = Table(displayName=display_name, ref=table_ref)
            style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tbl.tableStyleInfo = style
            ws.add_table(tbl)
            logger.info(f"Table angelegt: name={display_name}, ref={table_ref}")
        except Exception as e:
            logger.warning(f"Table konnte nicht angelegt werden (name={display_name}, ref={table_ref}): {e}")

        # Merges innerhalb der Tabelle (optional / informativ)
        merges_touching = t.get("merged_cells_in_table") or []
        if merges_touching:
            for m in merges_touching:
                rng = m.get("range")
                if rng:
                    try:
                        ws.merge_cells(rng)
                        logger.info(f"Merged (in table): {ws.title}!{rng}")
                    except Exception as e:
                        logger.warning(f"Merge übersprungen (in table {display_name}, {rng}): {e}")

def reconstruct_workbook(json_data: Dict[str, Any], out_xlsx: Path, logger: logging.Logger) -> None:
    wb = Workbook()
    # Default-Sheet löschen, wir legen passend neu an
    default_ws = wb.active
    wb.remove(default_ws)

    sheets: Dict[str, Any] = json_data.get("sheets") or {}
    if not sheets:
        logger.error("JSON enthält keine 'sheets'.")
        # dennoch leeres Workbook speichern
        wb.create_sheet("Sheet1")
        wb.save(out_xlsx)
        return

    first_visible_title = None

    for sheet_name, sheet_obj in sheets.items():
        try:
            ws = wb.create_sheet(title=sheet_name[:31] if sheet_name else "Sheet")
        except Exception:
            # Excel-Limit/ungültige Zeichen – fallback-Name
            base = (sheet_name or "Sheet").replace("/", "_").replace("\\", "_")
            ws = wb.create_sheet(title=base[:31])

        # Sichtbarkeit
        state = str(sheet_obj.get("sheet_state") or "visible")
        if state not in ("visible", "hidden", "veryHidden"):
            state = "visible"
        ws.sheet_state = state
        if state == "visible" and first_visible_title is None:
            first_visible_title = ws.title

        # Sheet-weite Merges (zusätzlich zu Table/Used-Range)
        sheet_merges = sheet_obj.get("merged_cells") or []

        # Tabellen?
        tables = sheet_obj.get("excel_tables")
        used_range = sheet_obj.get("used_range")

        if tables:
            write_tables(ws, tables, logger)
        elif used_range:
            write_used_range(ws, used_range, logger)
        else:
            # Nichts – trotzdem Merges anwenden, falls im JSON vorhanden
            logger.warning(f"Sheet '{ws.title}': weder 'excel_tables' noch 'used_range' vorhanden. Leeres Blatt erstellt.")

        # Sheet-Merges zuletzt anwenden (kann außerhalb von Tabellen liegen)
        write_merged_ranges(ws, sheet_merges, logger)

    # Aktives Blatt auf erstes sichtbares setzen
    if first_visible_title:
        wb.active = wb.sheetnames.index(first_visible_title)

    # Speichern
    try:
        wb.save(out_xlsx)
        logger.info(f"XLSX geschrieben: {out_xlsx.name}")
    except PermissionError as e:
        logger.error(f"Speichern fehlgeschlagen (PermissionError): {e}")
    except Exception as e:
        logger.error(f"Speichern fehlgeschlagen: {e}")

# ----------------------------
# Main
# ----------------------------

def main():
    base = script_dir()
    json_files = sorted([p for p in base.glob("*.json") if not p.name.startswith("~$")])

    if not json_files:
        print("Keine .json im Script-Verzeichnis gefunden.")
        return

    for jf in json_files:
        out_xlsx = base / f"{jf.stem}.xlsx"
        out_log = base / f"{jf.stem}.toxlsx.log.txt"
        logger = setup_logger(out_log)
        logger.info(f"Starte Rekonstruktion: {jf.name}")

        try:
            with jf.open("r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            logger.error(f"JSON konnte nicht gelesen werden: {e}")
            continue

        try:
            reconstruct_workbook(data, out_xlsx, logger)
        except Exception as e:
            logger.error(f"Fehler bei der Rekonstruktion: {e}", exc_info=True)

        logger.info("Fertig.")

if __name__ == "__main__":
    main()
