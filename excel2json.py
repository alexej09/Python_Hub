#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel -> JSON Strukturextraktion mit Logging und Merged-Cell-Behandlung.
- Liest alle .xlsx/.xlsm im Script-Verzeichnis
- Erzeugt je Datei eine {basename}.json und {basename}.log.txt
- Nutzt echte Excel-"Tables" (falls vorhanden); sonst Fallback auf genutzten Zellbereich
Autor: Dein KI-Kumpel :)
"""

from __future__ import annotations
import json
import logging
import sys
from pathlib import Path
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Dict, List, Tuple, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, range_boundaries

# ----------------------------
# Utils
# ----------------------------

def script_dir() -> Path:
    # Robust: funktioniert auch interaktiv
    try:
        return Path(__file__).resolve().parent
    except NameError:
        return Path.cwd().resolve()

def safe_filename_stem(p: Path) -> str:
    return p.stem

def json_default(obj: Any) -> Any:
    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    if isinstance(obj, Decimal):
        return float(obj)
    return str(obj)

def setup_logger(log_path: Path) -> logging.Logger:
    logger = logging.getLogger(log_path.stem)
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")
    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)
    # Zusätzlich: Konsolen-Ausgabe minimal
    sh = logging.StreamHandler(sys.stdout)
    sh.setFormatter(fmt)
    logger.addHandler(sh)
    return logger

def cell_addr(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"

# ----------------------------
# Merged Cells Handling
# ----------------------------

def merged_ranges_info(ws: Worksheet) -> List[Dict[str, Any]]:
    """Liste aller Merge-Ranges mit Anker und Value (Ankerwert)"""
    info = []
    for m in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = m.bounds
        anchor = cell_addr(min_row, min_col)
        anchor_value = ws.cell(row=min_row, column=min_col).value
        info.append({
            "range": str(m),
            "anchor": anchor,
            "rows": [min_row, max_row],
            "cols": [min_col, max_col],
            "value": anchor_value
        })
    return info

def build_merged_lookup(ws: Worksheet) -> Dict[Tuple[int, int], Tuple[int, int]]:
    """
    Liefert ein Lookup: (r,c) -> (anchor_r, anchor_c) wenn Zelle in einem Merge liegt.
    openpyxl gibt Werte nur in der Ankerzelle zurück; wir merken uns die Beziehung.
    """
    lookup = {}
    for m in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = m.bounds
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                lookup[(r, c)] = (min_row, min_col)
    return lookup

def get_value_with_merge(ws: Worksheet, r: int, c: int, merge_lu: Dict[Tuple[int, int], Tuple[int, int]]):
    """Wert der Zelle inkl. Merge-Ankerauflösung"""
    if (r, c) in merge_lu:
        ar, ac = merge_lu[(r, c)]
        return ws.cell(row=ar, column=ac).value
    return ws.cell(row=r, column=c).value

# ----------------------------
# Tabellen-Extraktion (echte Excel Tables)
# ----------------------------

def unique_headers(headers: List[Any], logger: logging.Logger) -> List[str]:
    seen = {}
    out = []
    for h in headers:
        base = str(h) if h is not None else ""
        if base not in seen:
            seen[base] = 1
            out.append(base)
        else:
            seen[base] += 1
            new_h = f"{base}_{seen[base]}"
            logger.warning(f"Duplizierter Header '{base}' → umbenannt zu '{new_h}'.")
            out.append(new_h)
    return out

def extract_excel_table(ws: Worksheet, table_obj, logger: logging.Logger) -> Dict[str, Any]:
    """
    Extrahiert Struktur & Daten einer Excel-Tabelle (openpyxl.worksheet.table.Table)
    """
    ref = table_obj.ref  # z.B. "A1:D20"
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    merge_lu = build_merged_lookup(ws)

    # Header-Zeile
    headers_raw = []
    for c in range(min_col, max_col + 1):
        headers_raw.append(get_value_with_merge(ws, min_row, c, merge_lu))
    headers = unique_headers(headers_raw, logger)

    # Datenzeilen
    rows = []
    for r in range(min_row + 1, max_row + 1):
        row_obj = {}
        for idx, c in enumerate(range(min_col, max_col + 1)):
            val = get_value_with_merge(ws, r, c, merge_lu)
            # JSON-freundliche Werte
            row_obj[headers[idx]] = val
            # Mapping ins Log
            logger.info(
                f"Map: {ws.title}!{cell_addr(r, c)} -> tables['{table_obj.displayName}'].rows[{len(rows)}]['{headers[idx]}']"
            )
        rows.append(row_obj)

    # Merged-Cells, die die Tabelle schneiden (zur Transparenz im JSON)
    merges_touching = []
    for m in ws.merged_cells.ranges:
        mc = m.bounds  # (min_col, min_row, max_col, max_row) – Achtung Reihenfolge!
        m_min_col, m_min_row, m_max_col, m_max_row = mc
        # einfache Überschneidungsprüfung
        if not (m_max_col < min_col or m_min_col > max_col or m_max_row < min_row or m_min_row > max_row):
            merges_touching.append({"range": str(m), "anchor": cell_addr(m_min_row, m_min_col)})

    return {
        "name": table_obj.displayName,
        "ref": ref,
        "headers": headers,
        "row_count": len(rows),
        "rows": rows,
        "merged_cells_in_table": merges_touching
    }

# ----------------------------
# Fallback: Struktur ohne Excel-Table (genutzter Bereich)
# ----------------------------

def extract_used_range(ws: Worksheet, logger: logging.Logger) -> Dict[str, Any]:
    """
    Wenn kein Excel-Table vorhanden ist: gesamten genutzten Bereich als Raster + Merges dokumentieren.
    """
    dim = ws.calculate_dimension()  # z.B. "A1:F20" (oder "A1:A1" bei leer)
    min_col, min_row, max_col, max_row = range_boundaries(dim)
    merge_lu = build_merged_lookup(ws)

    grid: List[List[Any]] = []
    for r in range(min_row, max_row + 1):
        row_vals = []
        for c in range(min_col, max_col + 1):
            v = get_value_with_merge(ws, r, c, merge_lu)
            row_vals.append(v)
            logger.info(
                f"Map: {ws.title}!{cell_addr(r, c)} -> used_range[{r - min_row}][{c - min_col}]"
            )
        grid.append(row_vals)

    return {
        "dimensions": dim,
        "row_count": len(grid),
        "col_count": max_col - min_col + 1,
        "grid": grid,
        "merged_cells": merged_ranges_info(ws)
    }

# ----------------------------
# Workbook-Verarbeitung
# ----------------------------

def process_workbook(xl_path: Path, out_json: Path, out_log: Path) -> None:
    logger = setup_logger(out_log)
    logger.info(f"Starte Verarbeitung: {xl_path.name}")

    try:
        wb = load_workbook(filename=xl_path, data_only=True, read_only=False, keep_links=True)
        logger.info(f"Workbook geladen. Sheets: {wb.sheetnames}")
    except Exception as e:
        logger.error(f"Workbook konnte nicht geladen werden: {e}")
        return

    result: Dict[str, Any] = {
        "file": xl_path.name,
        "sheets": {}
    }

    for ws in wb.worksheets:
        try:
            header = f"{xl_path.name}-{ws.title}"
            logger.info(f"--- Sheet: {ws.title} ---")
            sheet_obj: Dict[str, Any] = {
                "header": header,
                "sheet_state": ws.sheet_state,  # visible/hidden/veryHidden
            }

            # Dokumentiere Merges auf Sheet-Ebene
            merges = merged_ranges_info(ws)
            if merges:
                logger.info(f"Merged Ranges: {[m['range'] for m in merges]}")
            sheet_obj["merged_cells"] = merges

            # Excel-Tables prüfen
            tables = list(ws.tables.values())
            if tables:
                logger.info(f"Gefundene Excel-Tabellen: {[t.displayName for t in tables]}")
                sheet_tables = []
                for t in tables:
                    t_obj = extract_excel_table(ws, t, logger)
                    sheet_tables.append(t_obj)
                sheet_obj["excel_tables"] = sheet_tables
            else:
                logger.warning("Keine Excel-Tabellen gefunden. Fallback auf genutzten Bereich.")
                sheet_obj["used_range"] = extract_used_range(ws, logger)

            result["sheets"][ws.title] = sheet_obj

        except Exception as e:
            logger.error(f"Fehler im Sheet '{ws.title}': {e}", exc_info=True)

    try:
        with out_json.open("w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2, default=json_default)
        logger.info(f"JSON geschrieben: {out_json.name}")
    except Exception as e:
        logger.error(f"JSON konnte nicht geschrieben werden: {e}")

    logger.info("Verarbeitung abgeschlossen.")

# ----------------------------
# Main
# ----------------------------

def main():
    base = script_dir()

    excel_files = sorted([p for p in base.glob("*.xls*") if p.suffix.lower() in (".xlsx", ".xlsm") and not p.name.startswith("~$")])
    if not excel_files:
        print("Keine .xlsx/.xlsm im Script-Verzeichnis gefunden.")
        return

    for xl in excel_files:
        stem = safe_filename_stem(xl)
        out_json = base / f"{stem}.json"
        out_log = base / f"{stem}.log.txt"
        process_workbook(xl, out_json, out_log)

if __name__ == "__main__":
    main()
